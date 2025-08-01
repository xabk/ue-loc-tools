import re
from loguru import logger
from pathlib import Path
from dataclasses import dataclass, field
import pandas as pd  # TODO Get rid of pandas?
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

from libraries import (
    polib,  # Modified polib: _POFileParser.handle_oc only splits references by ', '
)
from libraries.utilities import LocTask, init_logging
from libraries.types import StringContextList


# -------------------------------------------------------------------------------------
# Defaults - These can be edited, only used if not overridden in configs
# (needed to make the script work standalone)
#
# Priority:
# 1. Script parameters in task list section of base.config.yaml (if task list provided)
# 2. Global params from base.config.yaml (if config file found and parameters found)
# 3. Defaults below (if no parameters found in config or no config found)
@dataclass
class ProcessTestAndHashLocales(LocTask):
    # TODO: Process all loc targets if none are specified
    # TODO: Change lambda to list to process all loc targets when implemented
    loc_targets: list = field(
        default_factory=lambda: ['Game']
    )  # Localization targets, empty = process all targets

    # Generate/update Excel file with all the strings from the specified loc targets
    external_context_xlsx: str | None = None
    external_context_targets: list[str] | None = None
    external_context_create_backup: bool = True

    # Additional context files, title: paths, relative to Content
    additional_context: dict[str, list[str]] | None = None
    add_context_fields: list[str] | None = None
    skip_other_fields: bool = False

    # TODO Remove and make part of external context
    # Localization targets from which to load string table references
    string_table_refs_targets: list | None = None
    # Narrative context (Excel → PO)
    narrative_context_file: str | None = None
    ### End of TODO

    # Debug ID/test locale is also a good source locale:
    # it's sorted, with debug IDs, repetition markers, asset names, and comments
    debug_ID_locale: str = 'io'
    hash_locale: str = 'ia-001'

    # Hash locale parameters
    hash_not_used_marker: str = 'NOT USED'  # If this is in comment, use not_used_prefix
    hash_prefix_not_used: str = '? '  # Prefix for strings with `not_used` in comment
    hash_prefix: str = '# '  # Prefix for each string in hash locale
    hash_suffix: str = ' ~'  # Suffix for each string in hash locale

    clear_translations: bool = False  # Start over? E.g., if ID length changed
    debug_prefix: str = '#'  # Prefix to use for debug ID, start over if changed
    debug_separator: str = ':'
    id_length: int = 4  # Num of digits in ID (#0001), start over if changed
    remove_source_loc_prefixes: list[str] | None = None

    encoding: str = 'utf-8-sig'  # PO file encoding
    sort_po: bool = True  # Sort the file by source reference?
    sort_by_key: bool = False  # Sort the file by key?

    # Delete leading/trailing whitespace in translations
    delete_unsafe_whitespace: bool = False
    # Create a CSV file with source, target, and context fields

    delete_comments_criteria: list = field(
        default_factory=lambda: [
            r'^Key:.*$',  # Delete 'Key: NNN' comments: we have them in msgctxt
        ]
    )
    # Delete occurences: we have them in 'SourceLocation: NNN' comments
    delete_occurrences: bool = True

    # Regex to match variables that we want to keep in 'translation'
    # TODO: Add support for UE/ICU syntax (plural, genders, etc.)
    # Looking for {variables} and <empty tags ... />
    var_regex: str = r'{[^}\[<]+}|<[^/>]+/>'

    comments_criteria: list = field(
        # list of rules, each rule is a list: [property to check, regex, comment to add]
        #  - property to check: msgid, msgctx, etc. See libraries/polib
        default_factory=lambda: [
            [  # Adding hints for strings with plurals
                'msgid',
                r'}\|plural\(',
                'Please adapt to your language plural rules. We only support keywords: '
                'zero, one, two, few, many, other.\n'
                'Check what these keywords mean: https://www.unicode.org/cldr/charts/47/supplemental/language_plural_rules.html.',
            ]
        ]
    )

    # Regex to match indices and make them zero-padded to fix the sorting
    ind_regex: str = r'([\[\(])([^\]\)]+)([\]\)])'  # Anything in () or []

    # Regex pattern to match IDs
    id_regex_pattern: str = r'^{prefix}(\d{{{id_length}}}):'

    # Skip strings with empty namespace in debug locales
    skip_empty_namespace: bool = False

    # Rules to label strings as Label: {label}
    string_label_rules: list[tuple[str, str, str]] | None = None

    boxdragon_filter_comments: bool = False

    # TODO: Do I need this here? Or rather in smth from uetools lib?
    content_dir: str = '../'

    # Actual regex based on id_length
    _id_regex: str | None = None
    _content_path: Path | None = None
    _external_context: dict[str | None, StringContextList] | None = None
    _external_context_fields: list[str] = field(
        default_factory=lambda: {
            'Key': {'Wrap': True, 'Width': 20},
            'Path': {'Wrap': False, 'Width': 30},
            'Source': {'Wrap': True, 'Width': 50},
            'Info': {'Wrap': True, 'Width': 30},
            'Context': {'Wrap': True, 'Width': 50},
            'MaxLength': {'Wrap': False, 'Width': 15},
            'Label': {'Wrap': False, 'Width': 20},
            'Screenshots': {'Wrap': False, 'Width': 15},
            'Timestamp': {'Wrap': False, 'Width': 20},
        }
    )

    # PO files, relative to Content directory
    _debug_id_file = 'Localization/{target}/{locale}/{target}.po'
    _hash_file = 'Localization/{target}/{locale}/{target}.po'

    # TODO Remove and make part of external context?
    _narrative_context_path: Path | None = None
    _source_ref_file = 'Localization/{target}/StringTableReferences.csv'
    _string_table_refs: dict[str, list[str]] | None = None
    _narrative_context: dict[str, str] | None = None
    ### End of TODO

    def post_update(self):
        super().post_update()
        self._content_path = Path(self.content_dir)
        if self.narrative_context_file:
            self._narrative_context_path = (
                self._content_path / self.narrative_context_file
            )
            if not (
                self._narrative_context_path.exists()
                and self._narrative_context_path.is_file()
            ):
                logger.warning(
                    'Narrative context file not found or not a file: '
                    f'{self._narrative_context_path}'
                )
                self._narrative_context_path = None

        self._id_regex = self.id_regex_pattern.format(
            prefix=re.escape(self.debug_prefix), id_length=self.id_length
        )

        if self.debug_ID_locale:
            self._debug_id_file = self._debug_id_file.format(
                target='{target}', locale=self.debug_ID_locale
            )
        if self.hash_locale:
            self._hash_file = self._hash_file.format(
                target='{target}', locale=self.hash_locale
            )

        self._string_table_refs = {}

    def load_external_data(self):
        if self.string_table_refs_targets:
            files = [
                self._source_ref_file.format(target=target)
                for target in self.string_table_refs_targets
            ]
            self._string_table_refs = self.load_string_table_refs(files)

        if self._narrative_context_path:
            self._narrative_context = self.load_narrative_context(
                [self._narrative_context_path]
            )

        if self.additional_context:
            self._external_context = self.load_external_context()

    def id_gen(
        self,
        number: int,
        id_length: int | None = None,
        prefix: str | None = None,
        separator: str | None = None,
        text: str | None = None,
        variables: list[str] | None = None,
    ) -> str:
        """
        Generate fixed-width #12345 IDs (number to use, optional ID width and prefix).
        """
        if not id_length:
            id_length = self.id_length
        if not prefix:
            prefix = self.debug_prefix
        id = prefix + str(number).zfill(id_length)
        if not separator:
            separator = self.debug_separator if self.debug_separator else ':'
        if text:
            id = f'{id}{separator}{text}'
        if variables:
            id = f'{id}{separator}{" ".join(variables)}'
        return id

    @staticmethod
    def ind_repl(match: re.Match, width: int = 5) -> str:
        """
        Generate a zero-padded (num) or [num] index.
        """
        index = re.sub(r'\d+', lambda match: match.group().zfill(width), match.group(2))
        return match.group(1) + index + match.group(3)

    def get_additional_comments(self, entry: polib.POEntry) -> list:
        """
        Get additional comments based on criteria
        """
        comments = []
        for [prop, crit, comment] in self.comments_criteria:
            if re.search(crit, getattr(entry, prop)):
                comments += [comment]
        return comments

    def should_delete_comment(self, comment: str) -> bool:
        for expr in self.delete_comments_criteria:
            if re.match(expr, comment):
                return True
        return False

    def load_string_table_refs_from_file(self, fpath: str) -> StringContextList:
        """
        Load string table references from CSV file for a single target

        Return a dict of StrTableName,Key : List of all references"""

        f_path = self._content_path / fpath
        if not f_path.exists() or not f_path.suffix == '.csv':
            return {}

        with open(f_path, mode='r', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            next(csv_reader)  # Skip headers
            string_table_refs = {}
            for row in csv_reader:
                (table_path, _, key) = row[0].rpartition(',')
                table_name = table_path.rpartition('/')[2]
                table_namespace = table_name.rpartition('.')[2]
                identity = f'{table_namespace},{key}'
                if identity not in string_table_refs:
                    string_table_refs[identity] = [row[1]]
                else:
                    string_table_refs[identity].append(row[1])

        return string_table_refs

    def load_string_table_refs(
        self, files: list[str] | None = None
    ) -> StringContextList:
        """
        Load string table references from CSV files for all targets

        Return a dict of StrTableName,Key : String with all references, one per line
        """

        data: StringContextList = {}

        if not files:
            return {}

        for file in files:
            if not data:
                data = self.load_string_table_refs_from_file(file)
                continue

            for key, value in self.load_string_table_refs_from_file(file).items():
                if key not in data:
                    data[key] = value
                else:
                    data[key].extend(value)
                    data[key] = list(set(data[key]))

        return data

    def get_references_for_key(self, key: str) -> list[str]:
        """
        Get all references for a key from a string tables
        """
        return self._string_table_refs.get(key, [])

    def load_narrative_context_from_file(self, fpath: str) -> StringContextList:
        """
        Load narrative context from Excel files
        """
        if not Path(fpath).exists():
            logger.warning(f'Narrative context file not found: {fpath}')
            return {}
        # Read all sheets with specific structure and avoid all conversions
        # Avoid any conversions, we want it all kept as is

        dfs = pd.read_excel(
            fpath,
            sheet_name=None,
            dtype=str,
            # keep_default_na=False,
            # na_values='',
            # converters=defaultdict(str),
        )

        narrative_context = {}

        for _, df in dfs.items():
            if 'Message ID' not in df.columns:
                continue

            context = False

            if 'General Context' in df.columns:
                df['Context'] = df['General Context']
                context = True
            if 'Direct Translation' in df.columns:
                df['Full message / Translated from Alienspeak'] = df[
                    'Direct Translation'
                ].where(df['Direct Translation'].notna(), df['Message Text'])
            else:
                df['Full message / Translated from Alienspeak'] = df['Message Text']

            df['Full message / Translated from Alienspeak'] = df[
                'Full message / Translated from Alienspeak'
            ].apply(
                lambda x: (
                    '\n'.join([s for s in x.splitlines() if s.strip()])
                    if isinstance(x, str)
                    else x
                )
            )

            df.dropna(subset=['Message ID'], inplace=True)

            columns = ['Context'] if context else []
            columns.append('Full message / Translated from Alienspeak')

            narrative_context.update(
                {
                    row['Message ID']: '\n'.join(
                        [
                            f'>>> {col}:\n{row[col]}'
                            for col in columns
                            if not pd.isna(row[col])
                        ]
                    )
                    for _, row in df.iterrows()
                }
            )

        narrative_context = {k: [v] for k, v in narrative_context.items() if v}

        return narrative_context

    def load_narrative_context(
        self, files: list[str] | None = None
    ) -> StringContextList:
        """
        Load string table references from CSV files for all targets

        Return a dict of StrTableName,Key : String with all references, one per line
        """

        narative_context: StringContextList = {}

        if not files:
            return {}

        for file in files:
            if not narative_context:
                narative_context = self.load_string_table_refs_from_file(file)
                continue

            for key, value in self.load_string_table_refs_from_file(file).items():
                if key not in narative_context:
                    narative_context[key] = value
                else:
                    narative_context[key].extend(value)
                    narative_context[key] = list(set(narative_context[key]))

        return narative_context

    def get_narrative_context_for_entry(self, entry: polib.POEntry) -> str | None:
        """
        Get narrative context for a key from a dictionary
        """
        # Extract asset name from the SourceLocation comment
        asset_name = None
        subtitle = None
        for comment in entry.comment.splitlines(False):
            if comment.startswith('SourceLocation:'):
                asset_name = re.search(r'/([^.]+)\.\1', comment)
                if asset_name:
                    asset_name = asset_name[1]

                subtitle = re.search(r'.mSubtitles\((\d+)\).', comment)
                if subtitle:
                    subtitle = int(subtitle[1])

                break

        context = self._narrative_context.get(asset_name, None)
        if context is None:
            return None

        context = '\n'.join(context)
        idx = None
        output_lines = []
        if subtitle is not None:
            for line in context.splitlines():
                if idx is not None:
                    idx += 1
                    if idx == subtitle:
                        line = '+++ ' + line
                if line.startswith('>>> Full message / Translated from Alienspeak'):
                    idx = -1
                output_lines.append(line)

            context = '\n'.join(output_lines)

        return context

    def load_context_from_csv(self, fpath: Path) -> StringContextList:
        """
        Load context from a CSV file
        """
        if not fpath.exists():
            logger.warning(f'Context file not found: {fpath}')
            return {}

        with open(fpath, mode='r', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            next(csv_reader)  # Skip headers
            data = {}
            for row in csv_reader:
                if row[0] not in data:
                    data[row[0]] = [row[1]]
                else:
                    data[row[0]].append(row[1])

        for key, value in data.items():
            data[key] = list(set(value))

        return data

    def load_context_from_xlsx(self, fpath: Path) -> StringContextList:
        """
        Load context from an Excel file
        """
        if not fpath.exists():
            logger.warning(f'Context file not found: {fpath}')
            return {}

        dfs = pd.read_excel(
            fpath,
            sheet_name=None,
            dtype=str,
            # keep_default_na=False,
            # na_values='',
            # converters=defaultdict(str),
        )

        data = {}

        for _, df in dfs.items():
            for _, row in df.iterrows():
                key = row[0]
                if key not in data:
                    data[key] = []

                for col in df.columns[1:]:
                    value = row[col]
                    if pd.notna(value):
                        data[key].append(f'{col}: {value}')

        for key, value in data.items():
            data[key] = list(set(value))

        return data

    def load_external_context_from_file(self, fpath: Path) -> StringContextList:
        """
        Load external context from a file
        """
        if fpath.suffix == '.csv':
            # Expecting references
            data = self.load_context_from_csv(fpath)
        elif fpath.suffix == '.xlsx':
            data = self.load_context_from_xlsx(fpath)
        else:
            logger.error(f'Unsupported file type: {fpath.suffix}')
            return {}

        if data is None:
            return {}

        return data

    def load_external_context(self) -> dict[str | None, StringContextList]:
        """
        Load external context from all files
        """
        if self.additional_context is None:
            return {}

        context = {}
        for title, files in self.additional_context.items():
            if not files:
                continue

            if not title:
                title = None

            context_cat = {}
            for file in files:
                data = self.load_external_context_from_file(self._content_path / file)
                for key, value in data.items():
                    if key not in context_cat:
                        context_cat[key] = value
                    else:
                        context_cat[key].extend(value)
                        context_cat[key] = list(set(context_cat[key]))

            context[title] = context_cat

        return context

    def get_external_context_for_key(self, key: str) -> list[str]:
        """
        Get external context for a key from a dictionary
        """
        context = []
        for title, data in self._external_context.items():
            if key in data:
                if title:
                    context += [f'{title}: {c}' for c in data[key]]
                else:
                    context += data[key]

        context = sorted(list(set(context)))

        if self.add_context_fields:
            ordered_context = []
            for field in self.add_context_fields:
                for ctx in context:
                    if ctx.startswith(f'{field}:'):
                        ordered_context.append(ctx)
            context = ordered_context
            if not self.skip_other_fields:
                context += [ctx for ctx in context if ctx not in ordered_context]

        return context

    def update_context_xlsx_file(self):
        if not self.external_context_xlsx or not self.external_context_targets:
            return

        # Read all PO files and create a DataFrame for each target
        data = {}
        for target in self.external_context_targets:
            po_entries = []
            po_file = self._content_path / self._debug_id_file.format(target=target)
            if not po_file.exists():
                logger.warning(f'PO file not found: {po_file}')
                continue

            po = polib.pofile(po_file, wrapwidth=0, encoding=self.encoding)
            for entry in po:
                loc_comments = []
                path = ''
                for comment in entry.comment.splitlines(False):
                    if re.match(r'^(Loc:\s|SourceLocation:\s)(.+)$', comment):
                        path = re.search(
                            r'(Loc:\s|SourceLocation:\s)(.+)$', comment
                        ).group(2)
                        continue

                    if comment not in loc_comments:
                        loc_comments.append(comment)

                loc_comments = '\n'.join(loc_comments)

                po_entries.append(
                    {
                        'Key': entry.msgctxt,
                        'Path': path,
                        'Source': entry.msgid,
                        'Info': loc_comments,
                    }
                )

            data[target] = po_entries

        # Read existing Excel file and update it with new data

        file_path = self._content_path / self.external_context_xlsx
        # Check if the file exists
        if file_path.exists():
            workbook = openpyxl.load_workbook(file_path)
            if self.external_context_create_backup:
                # Create a backup in ~ContextBackups
                backup_path = file_path.parent / '~ContextBackups'
                backup_path.mkdir(exist_ok=True)
                # Save the backup with a timestamp
                backup_path = (
                    backup_path / f'{datetime.now():%Y%m%d_%H%M%S}-{file_path.name}'
                )
                # Copy file path to backup path
                file_path.replace(backup_path)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove the default sheet

        for target, dataset in data.items():
            if target in workbook.sheetnames:
                sheet = workbook[target]
            else:
                sheet = workbook.create_sheet(title=target)

            # Define headers
            headers = {
                'Key': {'Wrap': True, 'Width': 50},
                'Path': {'Wrap': False, 'Width': 15},
                'Source': {'Wrap': True, 'Width': 50},
                'Info': {'Wrap': True, 'Width': 30},
                'Context': {'Wrap': True, 'Width': 50},
                'MaxLength': {'Wrap': False, 'Width': 15},
                'Label': {'Wrap': True, 'Width': 30},
                'Screenshot': {'Wrap': False, 'Width': 30},
                'Timestamp': {'Wrap': False, 'Width': 20},
            }

            # Write headers if they don't exist
            if sheet.max_row == 1:
                for col_num, (header, props) in enumerate(headers.items(), 1):
                    col_letter = get_column_letter(col_num)
                    cell = sheet[f'{col_letter}1']
                    cell.value = header
                    cell.font = Font(bold=True)
                # Freeze the first row
                sheet.freeze_panes = 'A2'

                # Set column widths and text wrapping
                for col_num, (header, props) in enumerate(headers.items(), 1):
                    col_letter = get_column_letter(col_num)
                    sheet.column_dimensions[col_letter].width = props['Width']
                    if props['Wrap']:
                        for row in range(1, sheet.max_row + 1):
                            cell = sheet[f'{col_letter}{row}']
                            cell.alignment = Alignment(wrap_text=True)

            # Create a dictionary to map headers to column letters
            header_to_col = {
                sheet[f'{get_column_letter(col_num)}1'].value: get_column_letter(
                    col_num
                )
                for col_num in range(1, sheet.max_column + 1)
            }

            # Add any missing headers at the end
            for header, props in headers.items():
                if header not in header_to_col:
                    col_num = sheet.max_column + 1
                    col_letter = get_column_letter(col_num)
                    cell = sheet[f'{col_letter}1']
                    cell.value = header
                    cell.font = Font(bold=True)
                    header_to_col[header] = col_letter
                    sheet.column_dimensions[col_letter].width = props['Width']
                    if props['Wrap']:
                        for row in range(1, sheet.max_row + 1):
                            cell = sheet[f'{col_letter}{row}']
                            cell.alignment = Alignment(wrap_text=True)

            # Create a dictionary to map keys to row numbers
            key_to_row = {
                sheet[f'{header_to_col["Key"]}{row}'].value: row
                for row in range(2, sheet.max_row + 1)
            }

            # Track existing keys
            existing_keys = set(key_to_row.keys())
            new_keys = set(row_data['Key'] for row_data in dataset)

            # Write or update data
            current_row = 2
            for row_data in dataset:
                key = row_data['Key']
                if key in key_to_row:
                    row_num = key_to_row[key]
                    current_row = row_num + 1
                    # Highlight changed cells
                    row_changed = False
                    for header in headers:
                        col_letter = header_to_col[header]
                        cell = sheet[f'{col_letter}{row_num}']
                        cell_value = row_data.get(header, None)
                        if cell_value is not None:
                            # Normalize newlines for comparison
                            cell_value_normalized = cell_value.replace(
                                '\r\n', '\n'
                            ).replace('\r', '\n')
                            cell_value_existing = (
                                (cell.value or '')
                                .replace('\r\n', '\n')
                                .replace('\r', '\n')
                            )
                            if cell_value_existing != cell_value_normalized:
                                cell.value = cell_value
                                cell.fill = PatternFill(
                                    start_color='FFFF99',
                                    end_color='FFFF99',
                                    fill_type='solid',
                                )
                                row_changed = True
                    if row_changed:
                        timestamp_cell = sheet[f'{header_to_col["Timestamp"]}{row_num}']
                        timestamp_cell.value = datetime.now().strftime(
                            '%Y-%m-%d %H:%M:%S'
                        )
                else:
                    sheet.insert_rows(current_row)
                    row_num = current_row
                    # Update key_to_row dictionary for subsequent rows
                    for existing_key in list(key_to_row.keys()):
                        if key_to_row[existing_key] >= current_row:
                            key_to_row[existing_key] += 1
                    key_to_row[key] = current_row
                    current_row += 1
                    # Highlight new row
                    for header in headers:
                        col_letter = header_to_col[header]
                        cell = sheet[f'{col_letter}{row_num}']
                        cell_value = row_data.get(header, None)
                        if cell_value is not None:
                            cell.value = cell_value
                        cell.fill = PatternFill(
                            start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'
                        )
                        if headers[header]['Wrap']:
                            cell.alignment = Alignment(wrap_text=True)
                    timestamp_cell = sheet[f'{header_to_col["Timestamp"]}{row_num}']
                    timestamp_cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Highlight rows that exist in old data but not in new data
            for key in existing_keys - new_keys:
                row_num = key_to_row[key]
                for header in headers:
                    col_letter = header_to_col[header]
                    cell = sheet[f'{col_letter}{row_num}']
                    cell.fill = PatternFill(
                        start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'
                    )
                timestamp_cell = sheet[f'{header_to_col["Timestamp"]}{row_num}']
                if not timestamp_cell.value:
                    timestamp_cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Save the workbook
        workbook.save(file_path)

    def find_max_ID(self) -> int:
        """
        Find max used debug ID in `targets` localization targets.
        Returns 0 if no debug IDs are used.
        """
        max_id = 0

        po_files = []
        missing_files = []

        for target in self.loc_targets:
            fname = self._content_path / self._debug_id_file.format(target=target)
            if not fname.exists():
                missing_files.append(fname)
            else:
                po_files.append(fname)

        if missing_files:
            logger.error(
                f'Check loc targets configuration. Missing PO files: {missing_files}'
            )

        if not po_files:
            logger.error('No PO files found at all. Max used ID is set to 0.')
            return 0

        for fname in po_files:
            po = polib.pofile(
                fname,
                wrapwidth=0,
                encoding=self.encoding,
            )
            local_max_id = max(
                [
                    int(re.search(self._id_regex, entry.msgstr).group(1))
                    for entry in po
                    if re.search(self._id_regex, entry.msgstr)
                ]
                + [0]
            )

            max_id = max(local_max_id, max_id)

        return max_id

    def process_debug_ID_locale(self, po_file: str, starting_id: int) -> int:
        """
        Process the PO file to insert #12345 IDs as 'translations'
        and add them to context
        """

        logger.info(f'Processing {po_file}')

        current_id = starting_id

        po = polib.pofile(po_file, wrapwidth=0, encoding=self.encoding)

        logger.info(
            'Source file translation rate: {rate:.0%}'.format(
                rate=po.percent_translated() / 100
            )
        )

        #
        # Make indices in source reference zero-padded to fix the sorting
        #
        # And sort the PO by source reference to overcome the 'randomness' of the default
        # Unreal Engine GUIDs and get at least some logical grouping
        #
        if self.sort_po:
            for entry in po:
                ctxt = re.sub(
                    r'\d+', lambda match: match.group().zfill(3), entry.msgctxt
                )
                occurrences = []
                for oc in entry.occurrences:
                    # Zero-pad array indices and add keys with zero-padded numbers to occurences
                    oc0 = ''
                    if '### Key: ' not in oc[0]:
                        oc0 = (
                            re.sub(self.ind_regex, self.ind_repl, oc[0])
                            + f' ### Key: {ctxt}'
                        )
                    else:
                        oc0 = re.sub(self.ind_regex, self.ind_repl, oc[0])
                    occurrences.append((oc0, oc[1]))
                entry.occurrences = occurrences

            po.sort()

        if self.sort_by_key:
            po.sort(
                key=lambda x: 'Z' * 100 + x.msgctxt
                if x.msgctxt.startswith(',')
                else x.msgctxt
            )

        # --- This is only needed for CSVs?
        # if self.delete_unsafe_whitespace:
        #     for entry in po:
        #         entry.msgid = entry.msgid.strip()

        if not self.clear_translations:
            # Check existing translations and log strings
            # that are translated but do not contain a debug ID
            odd_strings = [
                entry
                for entry in po.translated_entries()
                if not re.search(self._id_regex, entry.msgstr)
            ]

            if odd_strings:
                logger.warning(
                    f'Translated lines with no IDs in them ({len(odd_strings)}):'
                )
                for entry in odd_strings:
                    logger.warning(
                        '\n'.join([entry.msgctxt, entry.msgid, entry.msgstr])
                    )

        logger.info(f'Starting ID: {current_id}')

        strings = [entry.msgid for entry in po]

        # Iterate over all entries to generate debug IDs, patch and add comments
        for entry in po:
            variables = []

            # If we want to retranslate all entries or if entry is not translated
            if self.clear_translations or not entry.translated():
                variables = [
                    str(var) for var in re.findall(self.var_regex, entry.msgid)
                ]

                # Generate and save the ID
                entry.msgstr = self.id_gen(
                    number=current_id, text=entry.msgid, variables=variables
                )

                current_id += 1

            debug_ID = 'Debug ID:\t' + self.id_gen(
                number=current_id, variables=variables, separator=' '
            )

            asset_name = ''

            for comment in entry.comment.splitlines(False):
                if comment.startswith('SourceLocation:') or comment.startswith('Loc:'):
                    if asset_name := re.search(
                        r'/([^/]+?\.(cpp|h|csv))(\(\d+\))?$', comment
                    ):
                        asset_name = asset_name[1]
                    elif asset_name := re.search(r'/([^.]+)\.\1', comment):
                        asset_name = asset_name[1]
                    break

            if asset_name:
                debug_ID += f'\t\tAsset: {asset_name}'

            if strings.count(entry.msgid) > 1:
                debug_ID += '\t\t// ###Rep###'

            new_comments = []
            new_comments.append(debug_ID)
            for comment in entry.comment.splitlines(False):
                if comment.startswith('SourceLocation:'):
                    pattern = 'SourceLocation:\t'
                    if (
                        self.remove_source_loc_prefixes is not None
                        and len(self.remove_source_loc_prefixes) > 0
                    ):
                        pattern += '(' + '|'.join(self.remove_source_loc_prefixes) + ')'
                    comment = re.sub(pattern, 'Loc:\t', comment)
                if comment.startswith('Debug ID:'):
                    continue
                if self.should_delete_comment(comment):
                    continue
                if comment.startswith('Label:\t'):
                    continue
                if comment.startswith('InfoMetaData:\t'):
                    # Remove prefix, remove quotes around field name and value,
                    # unescape internal quotes
                    comment = comment.partition('InfoMetaData:\t')[2]
                    comment = re.sub(r'^"(.*?)" : "(.*?)"$', r'\1: \2', comment)
                    comment = comment.replace('\\"', '"')
                if comment not in new_comments:
                    new_comments.append(comment)

            if self._narrative_context:
                narrative_context = self.get_narrative_context_for_entry(entry)
                if narrative_context and narrative_context not in new_comments:
                    new_comments.append(narrative_context)

            for comment in self.get_additional_comments(entry):
                if comment not in new_comments:
                    new_comments.append(comment)

            for comment in self.get_external_context_for_key(entry.msgctxt):
                if comment not in new_comments:
                    new_comments.append(comment)

            # TODO: Use clean-up references code from string tables prep
            usage_references = self.get_references_for_key(entry.msgctxt)

            if usage_references:
                new_comments += ['Used in:']
                new_comments += usage_references

            if self.string_label_rules is not None:
                label_rules = {}
                for rule in self.string_label_rules:
                    field, pattern, label = rule
                    if label not in label_rules:
                        label_rules[label] = []
                    label_rules[label].append((field, pattern))

                added_labels = []
                for label in label_rules:
                    for field, pattern in label_rules[label]:
                        if re.search(pattern, getattr(entry, field)):
                            added_labels.append(f'Label: {label}')
                            break

                added_labels = '\t'.join(added_labels)

                if added_labels and added_labels not in new_comments:
                    new_comments.append(added_labels)

            entry.comment = (
                '\n'.join(new_comments).replace('\\n', '\n').replace('\\r', '')
            )

            if self.delete_occurrences:
                entry.occurrences = []

        # TODO: Check for duplicate IDs across all targets
        ids = [entry.msgstr for entry in po.translated_entries()]
        if len(set(ids)) != len(ids):
            logger.error(
                'Duplicate #IDs spotted, please recompile the debug IDs from scratch (use CLEAR_TRANSLATIONS = True)'
            )

        logger.info(
            'Target file translation rate: {rate:.0%}'.format(
                rate=po.percent_translated() / 100
            )
        )
        logger.info(f'Last used ID: {current_id - 1}')

        #
        # Save the file
        #
        po.save(po_file)
        logger.info(f'Saved target file: {po_file}')

        return current_id

    def process_hash_locale(self, po_file: str):
        """
        Open the PO, wrap every string in hash prefix and suffix, save the PO
        """
        po = polib.pofile(po_file, wrapwidth=0, encoding=self.encoding)
        logger.info(f'Opened hash locale file: {po_file}')

        for entry in po:
            namespace = str(entry.msgctxt).rpartition(',')[0]
            # If we want to skip empty namespace entries
            if namespace or not self.skip_empty_namespace:
                prefix = (
                    self.hash_prefix_not_used
                    if self.hash_not_used_marker in entry.comment
                    else self.hash_prefix
                )
                entry.msgstr = prefix + entry.msgid + self.hash_suffix

        po.save(po_file)
        logger.info(f'Saved target hash locale file: {po_file}')

        return True

    def process_locales(self):
        self.load_external_data()

        logger.info(f'Content path: {Path(self._content_path).absolute()}')

        logger.info(
            'String table references loaded: '
            f'{sum([len(r) for r in self._string_table_refs.values()])} '
            f'for {len(self._string_table_refs)} entries'
        )

        if self._narrative_context:
            logger.info(
                f'Narrative context loaded: {len(self._narrative_context)} entries'
            )

        if self._external_context:
            total_keys = sum(len(values) for values in self._external_context.values())
            total_values = sum(
                len(contexts)
                for items in self._external_context.values()
                for contexts in items.values()
            )

            logger.info(
                f'External context loaded. Categories: {len(self._external_context)} categories, '
                f'keys: {total_keys}, values: {total_values}'
            )

        starting_id = 1
        if not self.clear_translations and self.debug_ID_locale:
            starting_id = self.find_max_ID() + 1

        if not (self.debug_ID_locale or self.hash_locale):
            logger.error(
                'At least one of the locales must be specified (debug id or hash)'
            )
            return False

        hash_locales_processed = []
        debug_locales_processed = []
        errors = False

        for target in self.loc_targets:
            logger.info(f'Processing target: {target}')
            if self.debug_ID_locale:
                debug_id_PO = self._content_path / self._debug_id_file.format(
                    target=target
                )
                logger.info(f'Debug IDs PO file: {debug_id_PO}')
                if not debug_id_PO.exists():
                    logger.error(
                        f'Debug locale {self.debug_ID_locale} file not found: '
                        f'{debug_id_PO}'
                    )
                else:
                    starting_id = self.process_debug_ID_locale(
                        po_file=debug_id_PO,
                        starting_id=starting_id,
                    )
                    debug_locales_processed.append(target)

            if self.hash_locale:
                hash_loc_PO = self._content_path / self._hash_file.format(target=target)
                logger.info(f'Hash locale PO file: {hash_loc_PO}')
                logger.info(
                    'Hash symbols added '
                    f'({len(self.hash_prefix) + len(self.hash_suffix)}): '
                    f'`{self.hash_prefix}` and `{self.hash_suffix}`'
                )
                if not hash_loc_PO.exists():
                    logger.error(f'Hash locale file not found: {hash_loc_PO}')
                else:
                    self.process_hash_locale(hash_loc_PO)
                    hash_locales_processed.append(target)

        if self.debug_ID_locale and not len(debug_locales_processed) == len(
            self.loc_targets
        ):
            logger.error(
                'Not all debug locales have been processed: '
                f'{len(debug_locales_processed)} out of {len(self.loc_targets)}. '
                f'Loc targets: {self.loc_targets}. Debug locale: {self.debug_ID_locale}. '
                f'Processed debug locale targets: {hash_locales_processed}'
            )
            errors = True

        if self.hash_locale and not len(hash_locales_processed) == len(
            self.loc_targets
        ):
            logger.error(
                'Not all hash locales have been processed: '
                f'{len(hash_locales_processed)} out of {len(self.loc_targets)}. '
                f'Loc targets: {self.loc_targets}. Hash locale: {self.hash_locale}. '
                f'Processed hash locale targets: {hash_locales_processed}'
            )
            errors = True

        return errors

    def run(self) -> bool:
        """
        Run the locale processing task
        """
        result = self.process_locales()

        # Optional update of the context Excel file
        self.update_context_xlsx_file()

        return result


def main():
    init_logging()

    logger.info('')
    logger.info('--- Process debug id/test/source, and hash locales script start ---')
    logger.info('')

    task = ProcessTestAndHashLocales()

    task.read_config(Path(__file__).name)

    result = task.run()

    logger.info('')
    logger.info('--- Process debug id/test/source, and hash locales script end ---')
    logger.info('')

    if result:
        return 0

    return 1


# Process the files if the file isn't imported
if __name__ == '__main__':
    main()
