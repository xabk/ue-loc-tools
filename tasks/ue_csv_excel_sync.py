"""
CSV to Excel sync tool for Unreal Engine
Converts CSV files to Excel format and vice versa, with minimal external dependencies.
"""

import sys
import subprocess
import os
import csv
import argparse
from pathlib import Path
from enum import Enum
from dataclasses import dataclass, field
from typing import List, Optional
import importlib.util


# Try to import Unreal Engine module for path and source control utilities
try:
    import unreal

    UNREAL_AVAILABLE = True
except ImportError:
    UNREAL_AVAILABLE = False


def log(message: str, level: str = 'I') -> None:
    """Simple logging function for uniform logging throughout the script."""
    print(f'LogLocTools: [{level}] {message}')


if UNREAL_AVAILABLE:
    log('Unreal Engine module available - using UE path utilities')
else:
    log('Unreal Engine module not available - using relative paths')


# Check for and install openpyxl if needed
def ensure_openpyxl() -> bool:
    """Check if openpyxl is available, install if not."""
    spec = importlib.util.find_spec('openpyxl')
    if spec is not None:
        log('openpyxl is already available')
        return True
    else:
        log('openpyxl not found, attempting to install...')
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            log('openpyxl installed successfully')
            return True
        except subprocess.CalledProcessError as e:
            log(f'Failed to install openpyxl: {e}')
            return False


# Ensure openpyxl is available before importing it
if not ensure_openpyxl():
    log('Cannot proceed without openpyxl')
    sys.exit(1)

# Now we can import openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    log(f'Failed to import openpyxl: {e}')
    log('Please ensure openpyxl is installed correctly.')
    sys.exit(1)


# Configuration Constants - Easy to tweak
DEFAULT_CSV_DIR = 'StringTables'
DEFAULT_EXCEL_FILE = 'StringTables/!StringTables.xlsx'
DEFAULT_FALLBACK_CONTENT_DIR = '../../'

# Column width defaults
DEFAULT_COLUMN_WIDTHS = {
    'Key': 70,
    'SourceString': 70,
    'Context': 70,
    'MaxLength': 15,
}
DEFAULT_COLUMN_WIDTH = 30

# Excel styling defaults
DEFAULT_HEADER_FILL_COLOR = '222222'
DEFAULT_HEADER_FONT_COLOR = 'DDDDDD'

# File handling defaults
DEFAULT_CSV_ENCODING = 'utf-8-sig'
DEFAULT_EXCEL_SHEET_NAME_LIMIT = 31
DEFAULT_EXTRA_ROWS_FOR_STYLING = 1000

# Text processing settings
TEXT_NEWLINE_PATTERNS = ['\\r\\n', '\\n']  # Escaped newlines in CSV text
EXCEL_NEWLINE = '\n'  # Actual newline in Excel
CSV_NEWLINE = '\\r\\n'  # Escaped newline when writing to CSV
UNWANTED_ESCAPES = ["\\'", '\\"']  # Escapes to remove (quotes)

# Quote and newline escaping settings
ESCAPING_PAIRS = [
    ("'", "\\'"),  # Single quote: ' → \'
    ('"', '\\"'),  # Double quote: " → \"
]

# Column indices that should have quotes and newlines escaped (0-based)
# Typically SourceString is column 1, but this can be configured
COLUMNS_TO_ESCAPE_QUOTES = [1]  # SourceString column


class OperationMode(Enum):
    EXPORT = 'export'
    IMPORT = 'import'


@dataclass
class CSVExcelConverter:
    """Simple CSV to Excel converter for Unreal Engine."""

    csv_dir: str = DEFAULT_CSV_DIR
    excel_file: str = DEFAULT_EXCEL_FILE
    mode: OperationMode = OperationMode.EXPORT
    open_excel: bool = True  # Open Excel file with the converted CSV data
    force_overwrite: bool = False  # Force write to read-only files

    # Column width settings
    column_widths: dict[str, int] = field(
        default_factory=lambda: DEFAULT_COLUMN_WIDTHS.copy()
    )
    default_column_width: int = DEFAULT_COLUMN_WIDTH

    # Styling settings
    header_fill_color: str = DEFAULT_HEADER_FILL_COLOR
    header_font_color: str = DEFAULT_HEADER_FONT_COLOR

    # Text processing settings
    strip_whitespace_on_import: bool = False  # Strip whitespace in SourceStrings

    # Path settings - will be auto-detected if in UE, otherwise use fallback
    content_dir: str = field(default='', init=False)

    # Private fields
    _content_path: Path | None = field(default=None, init=False)
    _csv_path: Path | None = field(default=None, init=False)
    _excel_path: Path | None = field(default=None, init=False)

    def __post_init__(self):
        """Setup internal path variables after initialization."""
        self._setup_paths()

    def _setup_paths(self) -> None:
        """Setup internal path variables."""
        if UNREAL_AVAILABLE:
            try:
                # Get the project content directory
                content_dir = unreal.Paths.project_content_dir()

                self.content_dir = content_dir
                log(f'Using Unreal Engine content directory: {content_dir}')
            except Exception as e:
                log(f'Failed to get UE content path, using fallback: {e}', 'W')
                self.content_dir = DEFAULT_FALLBACK_CONTENT_DIR
        else:
            # Fallback for non-UE environments
            self.content_dir = DEFAULT_FALLBACK_CONTENT_DIR
            log(f'Using fallback content directory: {DEFAULT_FALLBACK_CONTENT_DIR}')

        self._content_path = Path(self.content_dir)
        self._csv_path = self._content_path / self.csv_dir
        self._excel_path = self._content_path / self.excel_file

    def _make_file_writable(self, file_path: Path, context: str = '') -> bool:
        """Make a file writable by changing its permissions."""
        try:
            import stat

            file_path.chmod(file_path.stat().st_mode | stat.S_IWRITE)
            context_str = f' ({context})' if context else ''
            log(f'Made file writable{context_str}: {file_path}')
            return True
        except Exception as e:
            log(f'Failed to make file writable: {file_path}: {e}', 'E')
            return False

    def _mark_file_for_add(self, file_path: Path) -> bool:
        """Mark a new file for add in source control if available.

        Args:
            file_path: Path to the file to mark for add

        Returns:
            True if successful or not needed, False if failed
        """
        if not UNREAL_AVAILABLE:
            return True  # No source control, no need to mark for add

        try:
            if (
                not unreal.SourceControl.is_available()
                or not unreal.SourceControl.is_enabled()
            ):
                log(f'Source control not available, skipping mark for add: {file_path}')
                return True  # No source control, no need to mark for add

            log(f'Marking file for add in source control: {file_path}')
            success = unreal.SourceControl.mark_file_for_add(
                str(file_path), silent=False
            )

            if success:
                log(f'Successfully marked file for add: {file_path}')
            else:
                log(f'Failed to mark file for add: {file_path}', 'W')
                error_msg = unreal.SourceControl.last_error_msg()
                if error_msg:
                    log(f'Source control error: {error_msg}', 'W')

            return success

        except Exception as e:
            log(f'Error marking file for add: {file_path}: {e}', 'W')
            return False

    def _checkout_file_if_needed(self, file_path: Path) -> bool:
        """Check out a file using Unreal's source control if available and needed."""

        # If file doesn't exist yet, mark it for add in source control (if available)
        if not file_path.exists():
            if UNREAL_AVAILABLE:
                try:
                    if (
                        unreal.SourceControl.is_available()
                        and unreal.SourceControl.is_enabled()
                    ):
                        # File will be created, so mark it for add after creation
                        log(f'New file will be marked for add: {file_path}')
                        # Note: We'll mark it for add after the file is actually created
                except Exception as e:
                    log(f'Could not check source control status: {e}', 'W')

            return True  # No need to do anytihng special, it's a new file

        if file_path.exists() and not UNREAL_AVAILABLE:
            if self.force_overwrite:
                return self._make_file_writable(file_path, 'force mode')
            else:
                # Check if file is already writable
                if self._is_file_writable(file_path):
                    log(f'File is already writable (no Unreal Engine): {file_path}')
                    return True
                else:
                    log(
                        f'File is not writable, Unreal not available, and force mode not enabled: {file_path}',
                        'W',
                    )
                    return False

        # File exists and we're in Unreal
        try:
            # Use the official unreal.SourceControl API
            # First check if source control is available and enabled
            if (
                not unreal.SourceControl.is_available()
                or not unreal.SourceControl.is_enabled()
            ):
                log('Source control is not available or not enabled')
                # Fall back to manual approach if force is enabled
                if self.force_overwrite:
                    return self._make_file_writable(
                        file_path, 'no source control, force mode'
                    )
                else:
                    log(
                        f'Source control not available and force mode not enabled: {file_path}',
                        'W',
                    )
                    return False

            # Query the file state first
            log(f'Querying source control state for: {file_path}')
            file_state = unreal.SourceControl.query_file_state(
                str(file_path), silent=True
            )

            if not file_state or not file_state.is_valid:
                log(f'Could not query file state for: {file_path}', 'W')
                # Fall back to manual approach if force is enabled
                if self.force_overwrite:
                    return self._make_file_writable(
                        file_path, 'invalid state, force mode'
                    )
                return False

            log(
                f'File state - Checked out: {file_state.is_checked_out}, Can edit: {file_state.can_edit}, Source controlled: {file_state.is_source_controlled}'
            )

            # If file is already checked out or can be edited, we're good
            if file_state.is_checked_out or file_state.can_edit:
                log(f'File is already editable: {file_path}')
                return True

            # If file is not under source control, make it writable
            if not file_state.is_source_controlled:
                log(f'File not under source control, making writable: {file_path}')
                return self._make_file_writable(file_path)

            # File is under source control but not checked out, try to check it out
            log(f'Attempting to check out file: {file_path}')
            checkout_success = unreal.SourceControl.check_out_file(
                str(file_path), silent=False
            )

            if checkout_success:
                # Verify the checkout was successful by re-querying
                new_state = unreal.SourceControl.query_file_state(
                    str(file_path), silent=True
                )
                if new_state and (new_state.is_checked_out or new_state.can_edit):
                    log(f'Successfully checked out file: {file_path}')
                    return True
                else:
                    log(
                        f'Checkout reported success but file is still not editable: {file_path}',
                        'W',
                    )
                    error_msg = unreal.SourceControl.last_error_msg()
                    if error_msg:
                        log(f'Source control error: {error_msg}', 'W')
            else:  # Checkout failed
                log(f'Failed to check out file: {file_path}', 'W')
                error_msg = unreal.SourceControl.last_error_msg()
                if error_msg:
                    log(f'Source control error: {error_msg}', 'W')

            # If checkout failed, try manual approach as last resort if force is enabled
            if self.force_overwrite:
                log(
                    f'Checkout failed, trying manual file permission change (force mode): {file_path}'
                )
                return self._make_file_writable(file_path, 'after checkout failure')
            else:
                log(f'Checkout failed and force mode not enabled: {file_path}', 'W')
                return False

        except Exception as e:
            log(f'Source control error for {file_path}: {e}', 'W')
            # Try manual approach as last resort if force is enabled
            if self.force_overwrite:
                return self._make_file_writable(file_path, 'fallback, force mode')
            else:
                log(
                    f'Source control error and force mode not enabled: {file_path}',
                    'W',
                )
                return False

    def _process_csv_to_excel_text(self, text: str) -> str:
        """Process text from CSV for Excel: convert escaped newlines to actual newlines, remove unwanted escapes."""
        if not isinstance(text, str):
            return text

        # Convert escaped newlines to actual newlines
        # Handle longer patterns first to avoid interference
        processed = text

        # Sort patterns by length (descending) to handle longer patterns first
        sorted_patterns = sorted(TEXT_NEWLINE_PATTERNS, key=len, reverse=True)
        for pattern in sorted_patterns:
            processed = processed.replace(pattern, EXCEL_NEWLINE)

        # Remove unwanted escapes
        for escape in UNWANTED_ESCAPES:
            processed = processed.replace(escape, escape[1:])  # Remove the backslash

        return processed

    def _process_excel_to_csv_text(self, text: str, should_escape: bool = False) -> str:
        """Process text from Excel for CSV: conditionally escape newlines and quotes."""
        if not isinstance(text, str):
            return text

        processed = text

        # Only escape newlines for columns that are configured to have escaping
        if should_escape:
            # Convert actual newlines to escaped format for CSV
            processed = processed.replace(
                '\r\n', CSV_NEWLINE
            )  # Handle Windows newlines first
            processed = processed.replace('\n', CSV_NEWLINE)  # Handle Unix newlines
            processed = processed.replace('\r', CSV_NEWLINE)  # Handle Mac newlines

            # Escape quotes for CSV format using configurable pairs
            for unescaped, escaped in ESCAPING_PAIRS:
                processed = processed.replace(unescaped, escaped)

        return processed

    def style_excel(self, sheet: Worksheet) -> None:
        """Apply styling to the Excel sheet."""
        header_fill = PatternFill(
            start_color=self.header_fill_color,
            end_color=self.header_fill_color,
            fill_type='solid',
        )
        header_font = Font(bold=True, color=self.header_font_color)
        alignment = Alignment(vertical='center', wrap_text=True)

        # Style headers
        for col_num, cell in enumerate(sheet[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            column_letter = get_column_letter(col_num)
            width = self.column_widths.get(cell.value, self.default_column_width)
            sheet.column_dimensions[column_letter].width = width

        # Apply wrap text to all cells
        for row in range(1, sheet.max_row + DEFAULT_EXTRA_ROWS_FOR_STYLING):
            for cell in sheet[row]:
                if cell.value is not None:
                    cell.alignment = alignment
                    cell.data_type = 's'
                    cell.number_format = '@'

        # Freeze first row
        sheet.freeze_panes = 'A2'

    def export_csvs_to_excel(self) -> bool:
        """Export CSV files to Excel format."""
        workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        if not self._csv_path.exists():
            log(f'CSV directory not found: {self._csv_path}', 'E')
            return False

        csv_files = list(self._csv_path.glob('*.csv'))
        if not csv_files:
            log(f'No CSV files found in: {self._csv_path}', 'E')
            return False

        for file_path in csv_files:
            try:
                with file_path.open('r', encoding=DEFAULT_CSV_ENCODING) as csv_file:
                    reader = csv.reader(csv_file)
                    sheet_name = file_path.stem[
                        :DEFAULT_EXCEL_SHEET_NAME_LIMIT
                    ]  # Excel sheet name limit

                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                    else:
                        sheet = workbook.create_sheet(title=sheet_name)

                    # Add headers
                    headers = next(reader, None)
                    if headers:
                        # Process headers
                        processed_headers = [
                            self._process_csv_to_excel_text(cell) for cell in headers
                        ]
                        sheet.append(processed_headers)

                    # Add data rows
                    for row in reader:
                        # Process each cell in the row
                        processed_row = [
                            self._process_csv_to_excel_text(cell) for cell in row
                        ]
                        sheet.append(processed_row)

                    self.style_excel(sheet)
                    log(f'Added {file_path.name} to Excel file')

            except Exception as e:
                log(f'Failed to process {file_path.name}: {e}', 'E')

        # Save workbook
        try:
            self._excel_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self._excel_path)
            log(f'Excel file created: {self._excel_path}')
            return True
        except Exception as e:
            log(f'Failed to save Excel file: {e}', 'E')
            return False

    def import_excel_to_csv(self) -> bool:
        """Import Excel sheets to CSV files."""
        if not self._excel_path.exists():
            log(f'Excel file not found: {self._excel_path}', 'E')
            return False

        try:
            workbook = load_workbook(self._excel_path)

            # Ensure CSV directory exists
            self._csv_path.mkdir(parents=True, exist_ok=True)

            for sheet_name in workbook.sheetnames:
                csv_file_path = self._csv_path / f'{sheet_name}.csv'
                sheet = workbook[sheet_name]

                try:
                    # Generate the new CSV data first (as list, not text)
                    raw_csv_data = self._generate_csv_data_from_sheet(sheet)

                    if not raw_csv_data:
                        log(f"Skipping empty sheet: '{sheet_name}'")
                        continue

                    # Check if CSV data has actually changed before validation (optimization)
                    if self._csv_data_are_identical(csv_file_path, raw_csv_data):
                        log(
                            f'CSV data unchanged, skipping validation and write: {csv_file_path}'
                        )
                        continue  # Skip this file entirely

                    # Data has changed, so validate and clean the string table data
                    try:
                        new_csv_data = self._validate_string_table_data(
                            raw_csv_data, sheet_name
                        )
                        log(f"String table validation passed for sheet: '{sheet_name}'")
                    except ValueError as validation_error:
                        log(
                            f"String table validation failed for sheet '{sheet_name}': {validation_error}",
                            'E',
                        )
                        continue  # Skip this sheet but continue with others

                    # Data has changed, attempt checkout
                    checkout_success = self._checkout_file_if_needed(csv_file_path)

                    # If checkout failed, warn but continue
                    if not checkout_success:
                        log(
                            f'Warning: Could not check out {csv_file_path}. Attempting to write anyway.',
                            'W',
                        )

                    # Write the CSV data to file
                    self._write_csv_data_to_file(new_csv_data, csv_file_path)

                    log(f"Exported sheet '{sheet_name}' to {csv_file_path}")

                except Exception as e:
                    log(f"Failed to export sheet '{sheet_name}': {e}", 'E')

            return True

        except Exception as e:
            log(f'Failed to load Excel file: {e}', 'E')
            return False

    def _is_file_writable(self, file_path: Path) -> bool:
        """Check if a file is writable without attempting to modify it."""

        if not file_path.exists():
            # File doesn't exist, so we can create it
            return True

        try:
            import stat

            file_stats = file_path.stat()
            return bool(file_stats.st_mode & stat.S_IWRITE)
        except Exception as e:
            log(f'Could not check file permissions for {file_path}: {e}', 'W')
            return False

    def _generate_csv_data_from_sheet(self, sheet: Worksheet) -> List[List[str]]:
        """Generate CSV data from an Excel sheet as a list of rows (not text)."""
        row_iterator = sheet.iter_rows(values_only=True)

        # Try to get the first row (header)
        try:
            header_row = next(row_iterator)
        except StopIteration:
            return []  # Empty sheet

        csv_data = []

        # Process header row (no escaping)
        clean_header = []
        for cell in header_row:
            if cell is not None:
                cell_text = str(cell)
                processed_text = self._process_excel_to_csv_text(
                    cell_text, should_escape=False
                )
                clean_header.append(processed_text)
            else:
                clean_header.append('')
        csv_data.append(clean_header)

        # Process data rows
        for row in row_iterator:
            clean_row = []
            for col_index, cell in enumerate(row):
                if cell is not None:
                    cell_text = str(cell)
                    # Check if this column should have quotes and newlines escaped
                    should_escape = col_index in COLUMNS_TO_ESCAPE_QUOTES
                    processed_text = self._process_excel_to_csv_text(
                        cell_text, should_escape=should_escape
                    )
                    clean_row.append(processed_text)
                else:
                    clean_row.append('')
            csv_data.append(clean_row)

        return csv_data

    def _validate_string_table_data(
        self, csv_data: List[List[str]], sheet_name: str
    ) -> List[List[str]]:
        """Validate string table data and return cleaned version."""

        log(f"Validating string table structure for sheet: '{sheet_name}'")

        # Check header row
        header_row = csv_data[0]
        if len(header_row) < 2:
            raise ValueError(
                f"Sheet '{sheet_name}' must have at least 2 columns (Key and SourceString)"
            )

        # Look for Key and SourceString columns
        key_col_index = None
        source_string_col_index = None

        for i, col_name in enumerate(header_row):
            col_name_lower = col_name.lower()
            if col_name_lower == 'key':
                key_col_index = i
            elif col_name_lower == 'sourcestring':
                source_string_col_index = i

        if key_col_index is None:
            raise ValueError(f"Sheet '{sheet_name}' must have a 'Key' column")
        if source_string_col_index is None:
            raise ValueError(f"Sheet '{sheet_name}' must have a 'SourceString' column")

        log(
            f'  Found Key column at index {key_col_index}, SourceString at index {source_string_col_index}'
        )

        # Process data rows and validate
        cleaned_data = [header_row]  # Keep header
        valid_entries_count = 0
        skipped_empty_rows = 0
        skipped_incomplete_rows = 0

        for row_index, row in enumerate(
            csv_data[1:], start=2
        ):  # Start from 2 for Excel row numbering
            # Ensure row has enough columns
            while len(row) < len(header_row):
                row.append('')

            key = row[key_col_index] if key_col_index < len(row) else ''

            if self.strip_whitespace_on_import:
                row[source_string_col_index] = row[source_string_col_index].strip()

            source_string = row[source_string_col_index]

            # Skip completely empty rows (allows for visual organization)
            if not any(cell for cell in row):
                skipped_empty_rows += 1
                continue

            # Validate non-empty key requirements
            if key:
                if not source_string:
                    raise ValueError(
                        f"Sheet '{sheet_name}', row {row_index}: Entry with key '{key}' has empty SourceString. All entries with keys must have text."
                    )
                valid_entries_count += 1
                cleaned_data.append(row)
            else:
                # Empty key - only allow if SourceString is also empty
                if source_string:
                    raise ValueError(
                        f"Sheet '{sheet_name}', row {row_index}: Entry has SourceString '{source_string}' but no Key. Entries must have both Key and SourceString or be empty."
                    )
                # Both key and source string are empty - skip this row
                skipped_incomplete_rows += 1

        # Check that we have at least one valid entry
        if valid_entries_count == 0:
            raise ValueError(
                f"Sheet '{sheet_name}' must have at least one entry with a non-empty Key and SourceString"
            )

        # Log validation summary
        log(
            f'  Validation summary: {valid_entries_count} valid entries, {skipped_empty_rows} empty rows skipped, {skipped_incomplete_rows} incomplete rows skipped'
        )
        return cleaned_data

    def _write_csv_data_to_file(
        self, csv_data: List[List[str]], file_path: Path
    ) -> None:
        """Write CSV data (list of rows) to a file with proper formatting."""

        # Check if this is a new file (for source control marking)
        is_new_file = not file_path.exists()

        with file_path.open('w', encoding=DEFAULT_CSV_ENCODING, newline='') as csv_file:
            if not csv_data:
                return  # Empty data

            # Write header without quotes
            header_writer = csv.writer(
                csv_file, quoting=csv.QUOTE_NONE, escapechar='\\'
            )
            header_writer.writerow(csv_data[0])

            # Write data rows (all fields quoted)
            if len(csv_data) > 1:
                data_writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
                for row in csv_data[1:]:
                    data_writer.writerow(row)

        # If this was a new file, mark it for add in source control
        if is_new_file:
            self._mark_file_for_add(file_path)

    def _csv_data_are_identical(
        self, file_path: Path, new_csv_data: List[List[str]]
    ) -> bool:
        """Compare existing CSV file data with new CSV data by parsing existing file."""

        if not file_path.exists():
            log(f'File does not exist, treating as different: {file_path}')
            return False  # File doesn't exist, so not identical

        try:
            # Parse existing file as CSV
            with file_path.open('r', encoding=DEFAULT_CSV_ENCODING) as existing_file:
                existing_reader = csv.reader(existing_file)
                existing_rows = list(existing_reader)

            # Compare the parsed CSV data directly with new data
            if existing_rows == new_csv_data:
                return True
            else:
                log(f'CSV data differs for: {file_path}')
                log(
                    f'  Existing rows: {len(existing_rows)}, New rows: {len(new_csv_data)}'
                )

                # Find first differing row
                min_rows = min(len(existing_rows), len(new_csv_data))
                for i in range(min_rows):
                    if existing_rows[i] != new_csv_data[i]:
                        log(f'  First difference at row {i}:')
                        log(f'    Existing: {existing_rows[i]}')
                        log(f'    New: {new_csv_data[i]}')
                        break
                else:
                    # All compared rows are identical, difference is in row count
                    if len(existing_rows) != len(new_csv_data):
                        log(
                            f'  Row count differs: existing={len(existing_rows)}, new={len(new_csv_data)}'
                        )

                return False

        except Exception as e:
            log(f'Could not parse CSV data for comparison in {file_path}: {e}', 'W')
            return False  # Assume different if we can't parse

    def run(self) -> bool:
        """Execute the conversion based on mode."""
        log(f'Starting CSV-Excel sync in {self.mode.value} mode')
        log(f'CSV directory: {self._csv_path}')
        log(f'Excel file: {self._excel_path}')

        if self.mode == OperationMode.EXPORT:
            success = self.export_csvs_to_excel()
            if success and self.open_excel:
                try:
                    os.startfile(str(self._excel_path))
                except Exception as e:
                    log(f'Could not open Excel file: {e}', 'W')

        elif self.mode == OperationMode.IMPORT:
            success = self.import_excel_to_csv()

        else:
            log(f'Invalid mode: {self.mode}', 'E')
            return False

        return success


def main() -> int:
    """Main function for standalone execution with command-line arguments."""
    parser = argparse.ArgumentParser(
        description='CSV to Excel sync tool for Unreal Engine',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s export                                    # Export CSV files to Excel
  %(prog)s import                                    # Import Excel sheets to CSV
  %(prog)s export --csv-dir Localization/Game       # Custom CSV directory
  %(prog)s export --excel-file StringTables.xlsx    # Custom Excel file
  %(prog)s export --content-dir /path/to/content    # Custom content directory
  %(prog)s export --no-open                         # Don't open Excel after export
  %(prog)s import --strip-whitespace                # Import with whitespace stripping enabled
        """,
    )

    parser.add_argument(
        'mode',
        choices=['export', 'import'],
        help='Operation mode: export CSV to Excel, or import Excel to CSV',
    )

    parser.add_argument(
        '--csv-dir',
        default=DEFAULT_CSV_DIR,
        help=f'Directory containing CSV files (default: {DEFAULT_CSV_DIR})',
    )

    parser.add_argument(
        '--excel-file',
        default=DEFAULT_EXCEL_FILE,
        help=f'Excel file path (default: {DEFAULT_EXCEL_FILE})',
    )

    parser.add_argument(
        '--content-dir',
        help=f'Content directory path (auto-detected if in UE, otherwise uses {DEFAULT_FALLBACK_CONTENT_DIR})',
    )

    parser.add_argument(
        '--no-open', action='store_true', help="Don't open Excel file after export"
    )

    parser.add_argument(
        '--force',
        action='store_true',
        help='Force write to read-only files (make them writable if source control checkout fails)',
    )

    parser.add_argument(
        '--strip-whitespace',
        action='store_true',
        help='Strip leading and trailing whitespace from cell values during import (default: False)',
    )

    args = parser.parse_args()

    # Create converter with parsed arguments
    converter = CSVExcelConverter()
    converter.csv_dir = args.csv_dir
    converter.excel_file = args.excel_file
    converter.mode = (
        OperationMode.EXPORT if args.mode == 'export' else OperationMode.IMPORT
    )
    converter.open_excel = not args.no_open
    converter.force_overwrite = args.force
    converter.strip_whitespace_on_import = args.strip_whitespace

    # Handle content directory
    if args.content_dir is not None:
        converter.content_dir = args.content_dir
        converter._setup_paths()
    # else: __post_init__ already handled auto-detection

    # Run the conversion
    success = converter.run()

    if success:
        log('Conversion completed successfully!')
        return 0
    else:
        log('Conversion failed!')
        return 1


# For Unreal Engine integration
def ue_export_csv_to_excel(
    csv_dir: str = DEFAULT_CSV_DIR,
    excel_file: str = DEFAULT_EXCEL_FILE,
    content_dir: Optional[str] = None,
    force_overwrite: bool = False,
    strip_whitespace_on_import: bool = False,
) -> bool:
    """Unreal Engine compatible function to export CSV to Excel."""
    converter = CSVExcelConverter()
    converter.csv_dir = csv_dir
    converter.excel_file = excel_file
    converter.force_overwrite = force_overwrite
    converter.strip_whitespace_on_import = strip_whitespace_on_import

    # If content_dir is explicitly provided, use it; otherwise auto-detect
    if content_dir is not None:
        converter.content_dir = content_dir
        converter._setup_paths()
    # else: __post_init__ already handled auto-detection

    converter.mode = OperationMode.EXPORT
    return converter.run()


def ue_import_excel_to_csv(
    csv_dir: str = DEFAULT_CSV_DIR,
    excel_file: str = DEFAULT_EXCEL_FILE,
    content_dir: Optional[str] = None,
    force_overwrite: bool = False,
    strip_whitespace_on_import: bool = False,
) -> bool:
    """Unreal Engine compatible function to import Excel to CSV."""
    converter = CSVExcelConverter()
    converter.csv_dir = csv_dir
    converter.excel_file = excel_file
    converter.force_overwrite = force_overwrite
    converter.strip_whitespace_on_import = strip_whitespace_on_import

    # If content_dir is explicitly provided, use it; otherwise auto-detect
    if content_dir is not None:
        converter.content_dir = content_dir
        converter._setup_paths()
    # else: __post_init__ already handled auto-detection

    converter.mode = OperationMode.IMPORT
    return converter.run()


if __name__ == '__main__':
    sys.exit(main())
